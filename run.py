import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# === CONFIG ===
folder_path = os.getcwd()  # Current folder where script runs
output_file = os.path.join(folder_path, "summary.xlsx")

# Create new workbook for the summary
summary_wb = Workbook()
summary_ws = summary_wb.active
summary_ws.title = "Summary"

# Header row
headers = ["Name", "Fase 1", "Fase 2", "Fase 3", "Total"]
summary_ws.append(headers)

# Make header bold
for cell in summary_ws[1]:
    cell.font = Font(bold=True)

# Temporary storage for all learners before sorting
learners_data = []

# Loop through all Excel files in the folder
for file_name in os.listdir(folder_path):
    if file_name.lower().endswith((".xlsx", ".xlsm")) and not file_name.startswith("~$") and file_name != "summary.xlsx":
        try:
            wb = load_workbook(file_name, data_only=True)
            if "Opsomming" not in wb.sheetnames:
                print(f"⚠ Skipping {file_name}: No 'Opsomming' sheet")
                continue
            
            ws = wb["Opsomming"]
            
            # Get marks
            fase1 = ws["E4"].value
            fase2 = ws["E5"].value
            fase3 = sum([
                ws["E6"].value or 0,
                ws["E7"].value or 0,
                ws["E8"].value or 0
            ])
            total = ws["E10"].value
            
            # Extract name from filename (remove ext and leading ".")
            name = os.path.splitext(file_name)[0].lstrip(".")
            
            # Store data for sorting later
            learners_data.append([name, fase1, fase2, fase3, total])
            
            print(f"✅ Processed: {name}")
        
        except Exception as e:
            print(f"❌ Error with {file_name}: {e}")

# Sort learners alphabetically by Name
learners_data.sort(key=lambda x: x[0].lower())

# Write sorted data to summary sheet
for row in learners_data:
    summary_ws.append(row)

# Auto-adjust column widths
for col in summary_ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get Excel column letter
    for cell in col:
        try:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        except:
            pass
    adjusted_width = max_length + 2
    summary_ws.column_dimensions[col_letter].width = adjusted_width

# Save summary file
summary_wb.save(output_file)
print(f"\n📄 Summary saved to {output_file}")
